import tkinter as tk
from tkinter import ttk, messagebox
import json
import os
from datetime import datetime, timedelta
import openpyxl
from openpyxl.cell.cell import MergedCell

CONFIG_FILE = "config.json"
DEFAULT_CONFIG = {
    "users": [],
    "current_user_index": -1,
    "station_info": {"name": "龙潭供电所", "county": "桃源县", "city": "常德市"},
    "rules": {
        "local": {"traffic": 0, "food": 40, "stay": 0, "misc": 0},
        "county": {"traffic": 0, "food": 0, "stay": 0, "misc_one_way": 15, "misc_round_trip": 30},
        "city": {"traffic": 0, "food": 0, "stay": 0, "misc_one_way": 25, "misc_round_trip": 50}
    },
    "template_paths": {
        "expense": "差旅费报销单模板.xlsx",
        "audit": "报销审核单模板.xlsx",
        "no_car": "未派车证明模板.xlsx"
    }
}

def num_to_cn_amount(num):
    if num == 0: return "零元整"
    units = ["", "拾", "佰", "仟"]
    big_units = ["", "万", "亿"]
    num_str = str(int(num))
    fraction = str(round(num - int(num), 2))[2:]
    result = ""
    length = len(num_str)
    for i, digit in enumerate(num_str):
        n = int(digit)
        if n != 0: result += "零壹贰叁肆伍陆柒捌玖"[n] + units[(length - 1 - i) % 4]
        if (length - 1 - i) % 4 == 0: result += big_units[(length - 1 - i) // 4]
    result = result.replace("零零", "零").strip("零")
    result += "元"
    if len(fraction) > 0:
        jiao = int(fraction[0])
        fen = int(fraction[1]) if len(fraction) > 1 else 0
        if jiao > 0: result += "零壹贰叁肆伍陆柒捌玖"[jiao] + "角"
        if fen > 0: result += "零壹贰叁肆伍陆柒捌玖"[fen] + "分"
    else: result += "整"
    return result

class TravelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("供电所差旅费工具 V2.6 (防崩+权限检测版)")
        self.root.geometry("960x780")
        self.config = self.load_config()
        self.trip_list = self.load_test_data() # 内置测试数据
        self.setup_ui()

    def load_test_data(self):
        trips = []
        # 预置7条测试数据，涵盖各种情况
        trips.append({"date": datetime(2024,5,6), "start": "龙潭", "end": "辖区", "food":40, "misc":0, "nocar":True, "reason":"线路巡视", "full_start_date":datetime(2024,5,6), "full_end_date":datetime(2024,5,6)})
        trips.append({"date": datetime(2024,5,8), "start": "龙潭", "end": "桃源县", "food":0, "misc":30, "nocar":False, "reason":"递交资料", "full_start_date":datetime(2024,5,8), "full_end_date":datetime(2024,5,8)})
        trips.append({"date": datetime(2024,5,10), "start": "龙潭", "end": "常德市", "food":0, "misc":25, "nocar":True, "reason":"技能培训", "full_start_date":datetime(2024,5,10), "full_end_date":datetime(2024,5,12)})
        trips.append({"date": datetime(2024,5,12), "start": "常德市", "end": "龙潭", "food":0, "misc":25, "nocar":False, "reason":"技能培训", "full_start_date":datetime(2024,5,10), "full_end_date":datetime(2024,5,12)})
        trips.append({"date": datetime(2024,5,15), "start": "龙潭", "end": "桃源县", "food":0, "misc":15, "nocar":False, "reason":"季度会议", "full_start_date":datetime(2024,5,15), "full_end_date":datetime(2024,5,16)})
        trips.append({"date": datetime(2024,5,16), "start": "桃源县", "end": "龙潭", "food":0, "misc":15, "nocar":False, "reason":"季度会议", "full_start_date":datetime(2024,5,15), "full_end_date":datetime(2024,5,16)})
        trips.append({"date": datetime(2024,5,20), "start": "龙潭", "end": "辖区", "food":40, "misc":0, "nocar":False, "reason":"临时抢修", "full_start_date":datetime(2024,5,20), "full_end_date":datetime(2024,5,20)})
        return trips

    def load_config(self):
        if not os.path.exists(CONFIG_FILE): return DEFAULT_CONFIG
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f: return json.load(f)
        except: return DEFAULT_CONFIG

    def save_config(self):
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(self.config, f, indent=4, ensure_ascii=False)

    # --- 核心：安全写入函数 (防崩版) ---
    def safe_write(self, ws, coord, value):
        try:
            # 1. 先检查该单元格是否为 MergedCell (只读部分)
            cell = ws[coord]
            if isinstance(cell, MergedCell):
                # 2. 如果是，尝试在所有合并区域中找到它的“父节点”
                found = False
                for rng in ws.merged_cells.ranges:
                    if coord in rng:
                        # 找到了父节点，写入父节点
                        ws.cell(row=rng.min_row, column=rng.min_col).value = value
                        found = True
                        break
                
                if not found:
                    # 3. 如果是 MergedCell 但找不到父节点 (极罕见，可能是 insert_rows 破坏了索引)
                    # 策略：跳过写入，打印警告，防止崩溃
                    print(f"Warning: Skipped writing to broken merged cell {coord}")
            else:
                # 4. 如果是普通单元格，直接写
                ws[coord] = value
        except Exception as e:
            # 5. 最后一道防线：任何写入错误都捕获，不让程序崩溃
            print(f"Error writing to {coord}: {str(e)}")

    # --- 核心：检查文件是否被占用 ---
    def check_file_lock(self, filename):
        if os.path.exists(filename):
            try:
                # 尝试以追加模式打开文件，如果被 Excel 占用会报错
                with open(filename, 'a'):
                    pass
            except PermissionError:
                return False
        return True

    def create_date_picker(self, parent):
        frame = ttk.Frame(parent)
        today = datetime.now()
        years = [str(y) for y in range(today.year - 1, today.year + 2)]
        months = [f"{m:02d}" for m in range(1, 13)]
        days = [f"{d:02d}" for d in range(1, 32)]
        cb_year = ttk.Combobox(frame, values=years, width=6, state="readonly")
        cb_year.set(today.year)
        cb_year.pack(side='left', padx=1)
        ttk.Label(frame, text="年").pack(side='left')
        cb_month = ttk.Combobox(frame, values=months, width=3, state="readonly")
        cb_month.set(f"{today.month:02d}")
        cb_month.pack(side='left', padx=1)
        ttk.Label(frame, text="月").pack(side='left')
        cb_day = ttk.Combobox(frame, values=days, width=3, state="readonly")
        cb_day.set(f"{today.day:02d}")
        cb_day.pack(side='left', padx=1)
        ttk.Label(frame, text="日").pack(side='left')
        return frame, cb_year, cb_month, cb_day

    def get_date_from_picker(self, picker_tuple):
        _, y, m, d = picker_tuple
        return f"{y.get()}-{m.get()}-{d.get()}"

    def set_picker_state(self, picker_tuple, state):
        _, y, m, d = picker_tuple
        y.config(state=state)
        m.config(state=state)
        d.config(state=state)

    def setup_ui(self):
        notebook = ttk.Notebook(self.root)
        notebook.pack(expand=True, fill='both')
        self.frame_gen = ttk.Frame(notebook)
        notebook.add(self.frame_gen, text="行程录入")
        self.setup_gen_tab()
        self.frame_user = ttk.Frame(notebook)
        notebook.add(self.frame_user, text="人员管理")
        self.setup_user_tab()
        self.frame_rules = ttk.Frame(notebook)
        notebook.add(self.frame_rules, text="设置")
        self.setup_rules_tab()
        self.refresh_trip_list_ui()

    def setup_gen_tab(self):
        left_panel = ttk.Frame(self.frame_gen, padding=10)
        left_panel.pack(side='left', fill='y')
        right_panel = ttk.Frame(self.frame_gen, padding=10)
        right_panel.pack(side='right', fill='both', expand=True)
        row = 0
        ttk.Label(left_panel, text="第一步：选择报销人").grid(row=row, column=0, columnspan=2, sticky='w')
        row+=1
        self.cb_users = ttk.Combobox(left_panel, state="readonly", width=25)
        self.cb_users.grid(row=row, column=0, columnspan=2, sticky='ew')
        self.update_user_combobox()
        row+=1
        ttk.Label(left_panel, text="第二步：录入行程").grid(row=row, column=0, columnspan=2, sticky='w', pady=10)
        row+=1
        ttk.Label(left_panel, text="出发日期:").grid(row=row, column=0, sticky='w')
        self.pk_start = self.create_date_picker(left_panel)
        self.pk_start[0].grid(row=row, column=1, sticky='w')
        row+=1
        ttk.Label(left_panel, text="起点:").grid(row=row, column=0, sticky='w')
        self.cb_start = ttk.Combobox(left_panel, values=["本所", self.config['station_info']['county'], self.config['station_info']['city']])
        self.cb_start.current(0)
        self.cb_start.grid(row=row, column=1, sticky='ew')
        row+=1
        ttk.Label(left_panel, text="终点:").grid(row=row, column=0, sticky='w')
        self.cb_end = ttk.Combobox(left_panel, values=["辖区线路", self.config['station_info']['county'], self.config['station_info']['city']])
        self.cb_end.bind("<<ComboboxSelected>>", self.on_end_point_change)
        self.cb_end.grid(row=row, column=1, sticky='ew')
        row+=1
        self.var_same_day = tk.BooleanVar(value=True)
        self.chk_same_day = ttk.Checkbutton(left_panel, text="当天往返", variable=self.var_same_day, command=self.on_sameday_change)
        self.chk_same_day.grid(row=row, column=1, sticky='w')
        row+=1
        ttk.Label(left_panel, text="返回日期:").grid(row=row, column=0, sticky='w')
        self.pk_end = self.create_date_picker(left_panel)
        self.pk_end[0].grid(row=row, column=1, sticky='w')
        self.set_picker_state(self.pk_end, "disabled")
        row+=1
        self.var_need_nocar = tk.BooleanVar(value=False)
        self.chk_nocar = ttk.Checkbutton(left_panel, text="需未派车证明", variable=self.var_need_nocar)
        self.chk_nocar.grid(row=row, column=0, sticky='w')
        ttk.Label(left_panel, text="事由:").grid(row=row, column=1, sticky='w')
        self.entry_reason = ttk.Entry(left_panel)
        self.entry_reason.insert(0, "差旅")
        self.entry_reason.grid(row=row+1, column=1, sticky='ew')
        row+=2
        ttk.Button(left_panel, text="⬇️ 添加到列表", command=self.add_trip_to_list).grid(row=row, column=0, columnspan=2, pady=15, sticky='ew')
        
        cols = ("日期", "地点", "金额", "未派车")
        self.tree_trips = ttk.Treeview(right_panel, columns=cols, show='headings', height=15)
        for c in cols: self.tree_trips.heading(c, text=c)
        self.tree_trips.column("日期", width=100); self.tree_trips.column("地点", width=200)
        self.tree_trips.column("金额", width=80); self.tree_trips.column("未派车", width=60)
        self.tree_trips.pack(fill='both', expand=True)
        
        btn_box = ttk.Frame(right_panel)
        btn_box.pack(fill='x', pady=5)
        ttk.Button(btn_box, text="删除选中行", command=self.del_trip_from_list).pack(side='left')
        ttk.Button(btn_box, text="清空列表", command=self.clear_trip_list).pack(side='left', padx=5)
        
        bottom_frame = ttk.LabelFrame(right_panel, text="生成设置")
        bottom_frame.pack(fill='x', pady=10)
        ttk.Label(bottom_frame, text="填报日期:").pack(side='left', padx=5)
        self.pk_fill = self.create_date_picker(bottom_frame)
        self.pk_fill[0].pack(side='left')
        ttk.Button(bottom_frame, text="🚀 生成文件", command=self.generate_all_files).pack(side='right', padx=10)
        self.lbl_total = ttk.Label(right_panel, text="当前总金额: 0 元")
        self.lbl_total.pack(anchor='e')

    def on_end_point_change(self, event):
        if self.cb_end.get() == "辖区线路":
            self.var_same_day.set(True)
            self.chk_same_day.config(state='disabled')
            self.set_picker_state(self.pk_end, "disabled")
        else:
            self.chk_same_day.config(state='normal')
            self.on_sameday_change()

    def on_sameday_change(self):
        if self.var_same_day.get(): self.set_picker_state(self.pk_end, "disabled")
        else: self.set_picker_state(self.pk_end, "readonly")

    def add_trip_to_list(self):
        try:
            start_date = datetime.strptime(self.get_date_from_picker(self.pk_start), "%Y-%m-%d")
            end_date = start_date if self.var_same_day.get() else datetime.strptime(self.get_date_from_picker(self.pk_end), "%Y-%m-%d")
        except: return messagebox.showerror("错误", "日期无效")
        
        start_place, end_place = self.cb_start.get(), self.cb_end.get()
        trips = []
        if end_place == "辖区线路":
            trips.append({"date": start_date, "start": self.config['station_info']['name'].replace("供电所",""), "end": "辖区", 
                          "food": self.config['rules']['local']['food'], "misc": self.config['rules']['local']['misc'], 
                          "nocar": self.var_need_nocar.get(), "reason": self.entry_reason.get(), "full_start_date": start_date, "full_end_date": end_date})
        else:
            rule = self.config['rules']['county'] if end_place == self.config['station_info']['county'] else self.config['rules']['city']
            clean_start = start_place.replace("本所", self.config['station_info']['name'].replace("供电所",""))
            if self.var_same_day.get():
                trips.append({"date": start_date, "start": clean_start, "end": end_place, "food": 0, "misc": rule['misc_round_trip'], 
                              "nocar": self.var_need_nocar.get(), "reason": self.entry_reason.get(), "full_start_date": start_date, "full_end_date": end_date})
            else:
                trips.append({"date": start_date, "start": clean_start, "end": end_place, "food": 0, "misc": rule['misc_one_way'], 
                              "nocar": self.var_need_nocar.get(), "reason": self.entry_reason.get(), "full_start_date": start_date, "full_end_date": end_date})
                trips.append({"date": end_date, "start": end_place, "end": clean_start, "food": 0, "misc": rule['misc_one_way'], 
                              "nocar": False, "reason": self.entry_reason.get(), "is_return_trip": True})
        
        for t in trips: self.trip_list.append(t)
        self.refresh_trip_list_ui()

    def del_trip_from_list(self):
        if self.tree_trips.selection():
            del self.trip_list[self.tree_trips.index(self.tree_trips.selection()[0])]
            self.refresh_trip_list_ui()

    def clear_trip_list(self):
        self.trip_list = []
        self.refresh_trip_list_ui()

    def refresh_trip_list_ui(self):
        for i in self.tree_trips.get_children(): self.tree_trips.delete(i)
        total = 0
        for t in self.trip_list:
            cost = t['food'] + t['misc']
            total += cost
            self.tree_trips.insert('', 'end', values=(t['date'].strftime("%m-%d"), f"{t['start']}->{t['end']}", cost, "是" if t.get('nocar') else "-"))
        self.lbl_total.config(text=f"当前总金额: {total} 元")

    def generate_all_files(self):
        if not self.trip_list: return messagebox.showerror("错误", "请先添加行程")
        if self.cb_users.current() == -1: return messagebox.showerror("错误", "请选择报销人")
        user = self.config['users'][self.cb_users.current()]
        try: fill_date = datetime.strptime(self.get_date_from_picker(self.pk_fill), "%Y-%m-%d")
        except: return messagebox.showerror("错误", "日期错误")

        self.trip_list.sort(key=lambda x: x['date'])
        total_money = sum([t['food'] + t['misc'] for t in self.trip_list])
        min_date, max_date = self.trip_list[0]['date'], self.trip_list[-1]['date']
        date_desc = f"自 {min_date.year} 年 {min_date.month} 月 {min_date.day} 日 至 {max_date.year} 年 {max_date.month} 月 {max_date.day} 日 计 {(max_date - min_date).days + 1} 天"
        file_suffix = f"{user['name']}_{fill_date.strftime('%m%d')}"

        # 检查文件占用
        f1_name = f"1_差旅费报销单_{file_suffix}.xlsx"
        f2_name = f"2_报销审核单_{file_suffix}.xlsx"
        if not self.check_file_lock(f1_name) or not self.check_file_lock(f2_name):
            return messagebox.showerror("错误", "生成的表格文件(如 1_差旅费...xlsx) 正被 Excel 打开。\n请先关闭这些文件，然后再点击生成！")

        try:
            wb = openpyxl.load_workbook(self.config['template_paths']['expense'])
            ws = wb.active
            self.safe_write(ws, 'K2', fill_date.year)
            self.safe_write(ws, 'M2', fill_date.month)
            self.safe_write(ws, 'O2', fill_date.day)
            self.safe_write(ws, 'B3', self.config['station_info']['name'])
            self.safe_write(ws, 'G3', self.config['station_info']['name'])
            self.safe_write(ws, 'B4', user['name'])
            self.safe_write(ws, 'E4', self.trip_list[0]['reason'])
            self.safe_write(ws, 'G4', "详见明细")
            self.safe_write(ws, 'J4', date_desc)
            
            curr_row = 8
            orig_rows = 6
            for i, t in enumerate(self.trip_list):
                if i >= orig_rows: ws.insert_rows(curr_row)
                self.safe_write(ws, f'A{curr_row}', t['date'].year)
                self.safe_write(ws, f'B{curr_row}', t['date'].month)
                self.safe_write(ws, f'C{curr_row}', t['date'].day)
                self.safe_write(ws, f'D{curr_row}', t['start'])
                self.safe_write(ws, f'E{curr_row}', t['end'])
                if t['food']: 
                    self.safe_write(ws, f'H{curr_row}', 1)
                    self.safe_write(ws, f'I{curr_row}', t['food'])
                if t['misc']: 
                    self.safe_write(ws, f'M{curr_row}', t['misc'])
                curr_row += 1
            
            r_tot, r_bk = 14 + max(0, len(self.trip_list) - orig_rows), 15 + max(0, len(self.trip_list) - orig_rows)
            
            self.safe_write(ws, f'G{r_tot}', num_to_cn_amount(total_money))
            self.safe_write(ws, f'C{r_bk}', user['name'])
            self.safe_write(ws, f'F{r_bk}', user['card'])
            self.safe_write(ws, f'K{r_bk}', user['bank'])
            self.safe_write(ws, f'N{r_bk}', user['phone'])
            
            wb.save(f1_name)

            wb2 = openpyxl.load_workbook(self.config['template_paths']['audit'])
            ws2 = wb2.active
            self.safe_write(ws2, 'K4', fill_date.year)
            self.safe_write(ws2, 'M4', fill_date.month)
            self.safe_write(ws2, 'O4', fill_date.day)
            self.safe_write(ws2, 'E6', self.config['station_info']['name'])
            self.safe_write(ws2, 'J10', total_money)
            self.safe_write(ws2, 'C11', num_to_cn_amount(total_money))
            self.safe_write(ws2, 'C12', user['name'])
            self.safe_write(ws2, 'F12', user['card'])
            self.safe_write(ws2, 'K12', user['bank'])
            self.safe_write(ws2, 'N12', user['phone'])
            wb2.save(f2_name)

            nocar_trips = [t for t in self.trip_list if t.get('nocar')]
            for t in nocar_trips:
                wb3 = openpyxl.load_workbook(self.config['template_paths']['no_car'])
                ws3 = wb3.active
                self.safe_write(ws3, 'F3', t['date'].year)
                self.safe_write(ws3, 'H3', t['date'].month)
                self.safe_write(ws3, 'J3', t['date'].day)
                self.safe_write(ws3, 'B5', self.config['station_info']['name'])
                self.safe_write(ws3, 'E5', user['name'])
                self.safe_write(ws3, 'H5', t['end'])
                self.safe_write(ws3, 'B7', t['reason'])
                fs, fe = t.get('full_start_date', t['date']), t.get('full_end_date', t['date'])
                self.safe_write(ws3, 'B8', fs.month)
                self.safe_write(ws3, 'D8', fs.day)
                self.safe_write(ws3, 'F8', fe.month)
                self.safe_write(ws3, 'H8', fe.day)
                wb3.save(f"3_未派车_{user['name']}_{fs.strftime('%m%d')}_至_{t['end']}.xlsx")

            messagebox.showinfo("成功", f"生成完毕！\n- 报销单: 1份\n- 审核单: 1份\n- 未派车证明: {len(nocar_trips)}份")

        except Exception as e:
            messagebox.showerror("运行出错", str(e))

    def setup_user_tab(self):
        p = ttk.Frame(self.frame_user, padding=10)
        p.pack(fill='both', expand=True)
        cols = ("姓名", "联系电话", "开户银行", "银行卡号")
        self.tree = ttk.Treeview(p, columns=cols, show='headings', height=10)
        for c in cols: self.tree.heading(c, text=c); self.tree.column(c, width=150)
        self.tree.pack(fill='x')
        frame_input = ttk.Frame(p)
        frame_input.pack(pady=10)
        self.entries_user = {}
        for i, col in enumerate(cols):
            ttk.Label(frame_input, text=col).grid(row=0, column=i, padx=5)
            if col == "开户银行":
                e = ttk.Combobox(frame_input, width=15, values=["中国农业银行", "中国工商银行", "中国建设银行", "中国邮政储蓄银行", "农村信用社", "长沙银行", "中国银行"])
            else: e = ttk.Entry(frame_input, width=15)
            e.grid(row=1, column=i, padx=5)
            self.entries_user[col] = e
        btn_box = ttk.Frame(p)
        btn_box.pack(pady=5)
        ttk.Button(btn_box, text="添加用户", command=self.add_user).pack(side='left', padx=5)
        ttk.Button(btn_box, text="删除选中", command=self.del_user).pack(side='left', padx=5)
        ttk.Button(btn_box, text="设为默认", command=self.set_default_user).pack(side='left', padx=5)
        self.refresh_user_list()
    
    def refresh_user_list(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        for u in self.config['users']: self.tree.insert('', 'end', values=(u['name'], u['phone'], u['bank'], u['card']))
    
    def update_user_combobox(self):
        names = [u['name'] for u in self.config['users']]
        self.cb_users['values'] = names
        if self.config['current_user_index'] >= 0 and self.config['current_user_index'] < len(names):
            self.cb_users.current(self.config['current_user_index'])

    def add_user(self):
        u = {k: v.get() for k, v in self.entries_user.items()}
        if not u["姓名"]: return
        self.config['users'].append({"name": u["姓名"], "phone": u["联系电话"], "bank": u["开户银行"], "card": u["银行卡号"]})
        self.save_config()
        self.refresh_user_list()
        self.update_user_combobox()
        for e in self.entries_user.values(): e.delete(0, tk.END)

    def del_user(self):
        if self.tree.selection():
            name = self.tree.item(self.tree.selection()[0])['values'][0]
            self.config['users'] = [u for u in self.config['users'] if u['name'] != name]
            self.config['current_user_index'] = -1
            self.save_config()
            self.refresh_user_list()
            self.update_user_combobox()

    def set_default_user(self):
        if self.cb_users.current() != -1:
            self.config['current_user_index'] = self.cb_users.current()
            self.save_config()
            messagebox.showinfo("成功", "已设为默认")

    def setup_rules_tab(self):
        p = ttk.Frame(self.frame_rules, padding=10)
        p.pack(fill='both', expand=True)
        grp_station = ttk.LabelFrame(p, text="基本信息")
        grp_station.pack(fill='x', pady=5)
        ttk.Label(grp_station, text="供电所名:").grid(row=0, column=0)
        self.entry_st_name = ttk.Entry(grp_station)
        self.entry_st_name.insert(0, self.config['station_info']['name'])
        self.entry_st_name.grid(row=0, column=1)
        ttk.Label(grp_station, text="所属县城:").grid(row=0, column=2)
        self.entry_st_county = ttk.Entry(grp_station)
        self.entry_st_county.insert(0, self.config['station_info']['county'])
        self.entry_st_county.grid(row=0, column=3)
        ttk.Label(grp_station, text="所属城市:").grid(row=0, column=4)
        self.entry_st_city = ttk.Entry(grp_station)
        self.entry_st_city.insert(0, self.config['station_info']['city'])
        self.entry_st_city.grid(row=0, column=5)
        grp_rule = ttk.LabelFrame(p, text="费用规则 (元)")
        grp_rule.pack(fill='x', pady=5)
        self.e_local_food = self.create_rule_entry(grp_rule, "[辖区内] 伙食:", 0, 0, 'local', 'food')
        self.e_county_round = self.create_rule_entry(grp_rule, "[县城] 往返杂费:", 1, 0, 'county', 'misc_round_trip')
        self.e_county_single = self.create_rule_entry(grp_rule, "[县城] 单程杂费:", 1, 2, 'county', 'misc_one_way')
        self.e_city_round = self.create_rule_entry(grp_rule, "[市区] 往返杂费:", 2, 0, 'city', 'misc_round_trip')
        self.e_city_single = self.create_rule_entry(grp_rule, "[市区] 单程杂费:", 2, 2, 'city', 'misc_one_way')
        ttk.Button(p, text="保存所有设置", command=self.save_all_settings).pack(pady=20)

    def create_rule_entry(self, parent, text, row, col, type, key):
        ttk.Label(parent, text=text).grid(row=row, column=col, pady=5)
        e = ttk.Entry(parent, width=8)
        e.insert(0, self.config['rules'][type][key])
        e.grid(row=row, column=col+1)
        return e
    
    def save_all_settings(self):
        self.config['station_info']['name'] = self.entry_st_name.get()
        self.config['station_info']['county'] = self.entry_st_county.get()
        self.config['station_info']['city'] = self.entry_st_city.get()
        try:
            self.config['rules']['local']['food'] = float(self.e_local_food.get())
            self.config['rules']['county']['misc_round_trip'] = float(self.e_county_round.get())
            self.config['rules']['county']['misc_one_way'] = float(self.e_county_single.get())
            self.config['rules']['city']['misc_round_trip'] = float(self.e_city_round.get())
            self.config['rules']['city']['misc_one_way'] = float(self.e_city_single.get())
        except ValueError: return messagebox.showerror("错误", "费用必须是数字")
        self.save_config()
        self.cb_start['values'] = ["本所", self.config['station_info']['county'], self.config['station_info']['city']]
        self.cb_end['values'] = ["辖区线路", self.config['station_info']['county'], self.config['station_info']['city']]
        messagebox.showinfo("成功", "设置已保存")

if __name__ == "__main__":
    root = tk.Tk()
    app = TravelApp(root)
    root.mainloop()
